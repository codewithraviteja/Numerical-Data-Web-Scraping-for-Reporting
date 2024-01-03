import csv
import pyperclip
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select

import random
import re

selenium.webdriver.support.select.Select
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait


PATH = "chromedriver.exe"
options = webdriver.ChromeOptions()
# options.experimental_options["debuggerAddress"] = "localhost:9014"
driver = webdriver.Chrome(PATH)
# driver = webdriver.Chrome(PATH, options=options)

driver.maximize_window()
i = 0
wb = openpyxl.load_workbook("data.xlsx")  # Replace with the actual Excel file name
sheet = wb.active
#########################################################################
# Get the maximum column index (number of columns in the sheet)
max_col_index = sheet.max_column

# Iterate through each column and print the data in the first row
for col_index in range(2, max_col_index + 1):
    yahoo_stock = sheet.cell(row=1, column=col_index).value
    trading_view_stock = sheet.cell(row=2, column=col_index).value
    print(f'Data in Column {chr(64 + col_index)}: {yahoo_stock} {trading_view_stock}')
##########################################################################   
# yahoo_stock = str(sheet.cell(row=1, column=2).value)
# trading_view_stock = str(sheet.cell(row=2, column=2).value)
######################################################

    print(trading_view_stock)

    ##########LOAD MORE#################
    # try:
    # 	for i in range(5):
    		
    # 		load_more = WebDriverWait(driver, 2).until(ec.visibility_of_element_located((By.XPATH, "//button[@data-overflow-tooltip-text='Load More ']")))
    # 		load_more.click()
    # 		print("worked")
    # 		time.sleep(3)
    # except:
    # 	print("Not found")

    # time.sleep(111111)
    ##########LOAD MORE#################
    url = "https://finance.yahoo.com/quote/" + yahoo_stock + ""
    driver.get(url)
    print(url)
    time.sleep(3)
    stock_name = WebDriverWait(driver, 60000).until(ec.visibility_of_element_located((By.XPATH, "//h1")))
    stock_name = stock_name.text
    print(stock_name)
    dividend_yield = WebDriverWait(driver, 60000).until(ec.visibility_of_element_located((By.XPATH, "//td[@data-test='DIVIDEND_AND_YIELD-value']")))
    get_yield = dividend_yield.text
    print(get_yield)
    average = WebDriverWait(driver, 60000).until(ec.visibility_of_element_located((By.XPATH, "//td[@data-test='ONE_YEAR_TARGET_PRICE-value']")))
    average = average.text
    print(average)
    driver.execute_script("window.scrollBy(0, 1000);")
    time.sleep(2)
    driver.execute_script("window.scrollBy(0, 500);")
    time.sleep(2)
    # driver.execute_script("window.scrollBy(0, 1500);")
    # time.sleep(2)		
    recommendation_rating = WebDriverWait(driver, 60000).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "[aria-label*='where 1 is Strong']")))
    recommendation_rating = recommendation_rating.text
    print(recommendation_rating)


    ## TRADING VIEW
    trading_view_get_page = str(sheet.cell(row=4, column=2).value)
    print(trading_view_get_page)
    trading_view_check = str(sheet.cell(row=4, column=1).value)
    print(trading_view_check)

    if trading_view_check == 'ALL':

        driver.get(trading_view_get_page)
        try:
          for i in range(5):
                
              load_more = WebDriverWait(driver, 2).until(ec.visibility_of_element_located((By.XPATH, "//button[@data-overflow-tooltip-text='Load More ']")))
              load_more.click()
              print("worked")
              time.sleep(3)
        except:
          print("Not found")    

    else:
        driver.get(trading_view_get_page)

    # MERCEDES-BENZ GROUP AG
    # get_stock_tarding_view = WebDriverWait(driver, 60000).until(ec.visibility_of_element_located((By.XPATH, "//a[contains(@title, 'MERCEDES-BENZ GROUP AG') and @href]")))
    # # get_stock_tarding_view = get_stock_tarding_view.text
    # # print(get_stock_tarding_view)
    time.sleep(2)
    # Find the matching <a> elements based on the XPath expression
    matching_elements = driver.find_elements(By.XPATH, f'//a[contains(@title, "{trading_view_stock}") and @href]')

    # Extract and print the href attribute values
    for element in matching_elements:
        href_value = element.get_attribute('href')
        symbol_text = href_value.split("symbols/")[1].rstrip('/')

        print("Symbol text:", symbol_text)

        a = symbol_text.replace('-', ':')
        print(a)




    get_stock_tarding_view = WebDriverWait(driver, 1).until(ec.visibility_of_element_located((By.XPATH, f'//tr[@data-rowkey="{a}"]/td[12]/div[contains(@class, "container")]')))
    # get_stock_tarding_view.click()
    get_stock_tarding_view = get_stock_tarding_view.text
    print(get_stock_tarding_view)

    driver.get(f'https://www.tradingview.com/symbols/{symbol_text}/forecast/')

    time.sleep(2)

    stock_name_tradingview = WebDriverWait(driver, 60000).until(ec.visibility_of_element_located((By.XPATH, "(//h1)")))
    stock_name_tradingview = stock_name_tradingview.text
    print(stock_name_tradingview)

    get_current_price = WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.XPATH, "(//span[contains(@class, 'symbol-last')])[1]")))
    get_current_price = get_current_price.text
    print(get_current_price)

    get_price_target = WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.XPATH, "//span[contains(@class, 'price-')]")))
    get_price_target = get_price_target.text
    print(get_price_target)

    get_common_data = WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.XPATH, "(//div[contains(@class, 'sectionSubtitle')])[1]")))
    get_common_data = get_common_data.text
    print(get_common_data)


    # Define patterns to match numbers with decimals
    pattern = re.compile(r'\d+\.\d+')

    # Find all matches in the text
    matches = pattern.findall(get_common_data)

    # Print the results
    if len(matches) >= 2:
        max_estimate = float(matches[0])
        min_estimate = float(matches[1])

        print("Max Estimate:", max_estimate)
        print("Min Estimate:", min_estimate)
    else:
        print("Unable to find both max and min estimates in the text.")

    def write_to_excel(file_path, *data):
        try:
            # Try to open the workbook
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            # If the workbook doesn't exist, create a new one
            workbook = Workbook()

        # Select the active sheet (create a new one if necessary)
        sheet = workbook.active

        # Find the next available row
        next_row = sheet.max_row + 1

        # Write data to the sheet
        for col_num, value in enumerate(data, start=1):
            sheet.cell(row=next_row, column=col_num, value=value)

        # Save the workbook
        workbook.save(file_path)

    current_date = datetime.now()
    formatted_date = current_date.strftime("%Y-%m-%d")#current_date.strftime("%B %d, %Y")
    print("Formatted date:", formatted_date)
    # Example usage
    file_path = 'report.xlsx'
    write_to_excel(file_path, formatted_date, "Yahoo Reports", stock_name, get_yield, average, recommendation_rating, "TradingView", trading_view_stock, get_current_price, get_price_target, max_estimate, min_estimate, get_stock_tarding_view)


driver.close()